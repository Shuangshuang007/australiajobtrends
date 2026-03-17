"""
Build pages_au/*.md from JSA Occupation profiles (australia/jds) for AI Exposure scoring.

Reads:
  - data_abs/occupations_combined.csv (title, anzsco_code)
  - ../australia/jds/Occupation profiles data*.xlsx  Table_2 (Description), Table_3 (Tasks)

Writes pages_au/{slug}.md (one per 4-digit ANZSCO). Slug = slugify(title) to match occupations_au.csv.

Run from Job Outlook Australia folder:
  python scripts/build_pages_au.py

Then run score_au.py to score each occupation.
"""

import csv
import os
import re
import glob

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.dirname(SCRIPT_DIR)
DATA_ABS = os.path.join(BASE_DIR, "data_abs")
COMBINED_CSV = os.path.join(DATA_ABS, "occupations_combined.csv")
AUSTRALIA_DIR = os.path.join(os.path.dirname(BASE_DIR), "australia")
JDS_DIR = os.path.join(AUSTRALIA_DIR, "jds")
PAGES_AU = os.path.join(BASE_DIR, "pages_au")


def slugify(title: str) -> str:
    """Same as make_csv_au: lowercase, spaces and punctuation to hyphens."""
    s = re.sub(r"[^\w\s-]", "", title.lower())
    s = re.sub(r"[-\s]+", "-", s).strip("-")
    return s or "occupation"


def norm_code(s) -> str:
    """4-digit ANZSCO unit group code."""
    if s is None:
        return ""
    digits = re.sub(r"\D", "", str(s))
    return digits[:4] if len(digits) >= 4 else digits or ""


def main():
    if not os.path.exists(COMBINED_CSV):
        print("Run build_occupations_from_anzsco.py first to create occupations_combined.csv")
        return

    # Occupation list: anzsco_code -> title, slug
    by_code = {}
    with open(COMBINED_CSV, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            code = norm_code(row.get("anzsco_code", ""))
            if len(code) != 4:
                continue
            title = (row.get("title") or "").strip()
            if not title:
                continue
            by_code[code] = {"title": title, "slug": slugify(title)}

    # JDS Occupation profiles
    pattern = os.path.join(JDS_DIR, "Occupation profiles data*.xlsx")
    files = sorted(glob.glob(pattern))
    if not files:
        print(f"No JDS file found at {pattern}")
        return

    try:
        import openpyxl
    except ImportError:
        print("pip install openpyxl")
        return

    wb = openpyxl.load_workbook(files[-1], read_only=True, data_only=True)
    if "Table_2" not in wb.sheetnames or "Table_3" not in wb.sheetnames:
        wb.close()
        print("JDS file missing Table_2 or Table_3")
        return

    # Table_2: ANZSCO Code, Occupation, Description (row 7 = header, 8+ = data)
    desc_by_code = {}
    for row in wb["Table_2"].iter_rows(min_row=8, values_only=True):
        if not row or len(row) < 3:
            continue
        code = norm_code(row[0])
        if len(code) != 4:
            continue
        desc = (row[2] or "").strip()
        if desc:
            desc_by_code[code] = desc

    # Table_3: ANZSCO Code, Occupation, Tasks (multiple rows per code)
    tasks_by_code = {}
    for row in wb["Table_3"].iter_rows(min_row=8, values_only=True):
        if not row or len(row) < 3:
            continue
        code = norm_code(row[0])
        if len(code) != 4:
            continue
        task = (row[2] or "").strip()
        if not task:
            continue
        if code not in tasks_by_code:
            tasks_by_code[code] = []
        tasks_by_code[code].append(task)

    wb.close()

    os.makedirs(PAGES_AU, exist_ok=True)
    written = 0
    for code, meta in by_code.items():
        title = meta["title"]
        slug = meta["slug"]
        desc = desc_by_code.get(code, "")
        tasks = tasks_by_code.get(code, [])

        md = [f"# {title}", ""]
        if desc:
            md.append("## Description")
            md.append("")
            md.append(desc)
            md.append("")
        if tasks:
            md.append("## Tasks")
            md.append("")
            for t in tasks:
                md.append(f"- {t}")
            md.append("")

        path = os.path.join(PAGES_AU, f"{slug}.md")
        with open(path, "w", encoding="utf-8") as f:
            f.write("\n".join(md))
        written += 1

    print(f"Wrote {written} markdown files to {PAGES_AU}")
    print("Next: python score_au.py")


if __name__ == "__main__":
    main()
