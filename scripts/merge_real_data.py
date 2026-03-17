"""
Merge real JSA and ABS data into data_abs/occupations_combined.csv.

Only overwrites num_jobs_2024, median_pay_annual, outlook_pct, entry_education
where real data exists; no fabricated values. See README_REAL_DATA.md for
download links and expected file/column names.

Expected files (optional):
  - data_abs/jsa_employment_projections.xlsx  (JSA: occupation, employment, growth %)
  - data_abs/abs_employment_by_occupation.xlsx or .csv  (ABS 6291 EQ08: employment by ANZSCO unit group)
  - data_abs/abs_earnings_by_occupation.xlsx or .csv  (ABS: occupation, median earnings)
  - ../australia/OSCA Category Descriptions.xlsx  (OSCA: fill entry_education from Skill Level by unit group)
  - ../australia/jds/Occupation profiles data*.xlsx  (JSA: Employed, Median weekly earnings, Annual employment growth by ANZSCO)

Column name hints (script looks for these in headers, case-insensitive):
  ANZSCO / Occupation code / Code  -> 4-digit unit group
  Employment / Employed / Jobs     -> num_jobs_2024
  Growth / Growth_pct / Projected growth  -> outlook_pct
  Median_weekly / Weekly earnings  -> convert to annual (* 52)
  Median_annual / Annual earnings   -> median_pay_annual
  Skill level / Education / AQF    -> entry_education (if present)

Usage (from Job Outlook Australia folder):
  python scripts/merge_real_data.py
"""

import csv
import json
import os
import re

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.dirname(SCRIPT_DIR)
DATA_ABS = os.path.join(BASE_DIR, "data_abs")
AUSTRALIA_DIR = os.path.join(os.path.dirname(BASE_DIR), "australia")
COMBINED_CSV = os.path.join(DATA_ABS, "occupations_combined.csv")
JSA_XLSX = os.path.join(DATA_ABS, "jsa_employment_projections.xlsx")
ABS_EMPLOYMENT = os.path.join(DATA_ABS, "abs_employment_by_occupation.xlsx")
ABS_EMPLOYMENT_CSV = os.path.join(DATA_ABS, "abs_employment_by_occupation.csv")
ABS_EARNINGS = os.path.join(DATA_ABS, "abs_earnings_by_occupation.xlsx")
ABS_EARNINGS_CSV = os.path.join(DATA_ABS, "abs_earnings_by_occupation.csv")
OSCA_DESCRIPTIONS = os.path.join(AUSTRALIA_DIR, "OSCA Category Descriptions.xlsx")
JDS_DIR = os.path.join(AUSTRALIA_DIR, "jds")
# Employment Projections (May 2025 - May 2035); Table_6 = 4-digit ANZSCO, 10-year % in col L
JDS_PROJECTIONS_DIR = os.path.join(AUSTRALIA_DIR, "JDS", "projections")

# ANZSCO Skill Level 1-5 -> AQF (entry_education), aligned with US education tiers
SKILL_LEVEL_TO_AQF = {
    1: "Bachelor degree or higher",
    2: "Diploma or Advanced Diploma",
    3: "Certificate III or IV",
    4: "Certificate II or III",
    5: "Certificate I or no formal qualification",
}


def slugify(title):
    """Same as make_csv_au: for matching imputed.json by slug."""
    s = re.sub(r"[^\w\s-]", "", (title or "").lower())
    s = re.sub(r"[-\s]+", "-", s).strip("-")
    return s or "occupation"


def norm_code(s):
    """Extract 4-digit ANZSCO unit group code."""
    if s is None:
        return ""
    s = str(s).strip()
    # Take first 4 digits if longer
    digits = re.sub(r"\D", "", s)
    return digits[:4] if len(digits) >= 4 else digits or ""


def find_column(header_row, hints):
    """Return 0-based column index where header matches any of hints."""
    for i, cell in enumerate(header_row):
        v = (str(cell).strip().lower() if cell is not None else "")
        for h in hints:
            if h in v:
                return i
    return -1


def merge_jsa(rows_by_code):
    """Merge JSA xlsx if present. Uses openpyxl."""
    if not os.path.exists(JSA_XLSX):
        return 0
    try:
        import openpyxl
    except ImportError:
        print("Install openpyxl to merge JSA xlsx: pip install openpyxl")
        return 0
    wb = openpyxl.load_workbook(JSA_XLSX, read_only=True, data_only=True)
    merged = 0
    for sheet in wb.worksheets:
        rows_iter = sheet.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            continue
        code_col = find_column(header, ["anzsco", "occupation code", "code", "unit group"])
        emp_col = find_column(header, ["employment", "employed", "jobs", "base"])
        growth_col = find_column(header, ["growth", "projected", "outlook", "change %"])
        if code_col < 0:
            continue
        for row in rows_iter:
            code = norm_code(row[code_col] if code_col < len(row) else None)
            if not code or code not in rows_by_code:
                continue
            r = rows_by_code[code]
            if emp_col >= 0 and emp_col < len(row) and row[emp_col] is not None:
                try:
                    r["num_jobs_2024"] = str(int(float(row[emp_col])))
                    merged += 1
                except (TypeError, ValueError):
                    pass
            if growth_col >= 0 and growth_col < len(row) and row[growth_col] is not None:
                try:
                    r["outlook_pct"] = str(int(round(float(row[growth_col]))))
                    merged += 1
                except (TypeError, ValueError):
                    pass
    wb.close()
    return merged


def merge_abs_employment(rows_by_code):
    """Merge ABS 6291 EQ08-style employment file (xlsx or csv) if present. Fills num_jobs_2024."""
    path = ABS_EMPLOYMENT_CSV if os.path.exists(ABS_EMPLOYMENT_CSV) else (ABS_EMPLOYMENT if os.path.exists(ABS_EMPLOYMENT) else None)
    if not path:
        return 0
    merged = 0
    if path.endswith(".csv"):
        with open(path, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                code = norm_code(row.get("ANZSCO") or row.get("Occupation code") or row.get("Code") or "")
                if not code or code not in rows_by_code:
                    continue
                r = rows_by_code[code]
                val = row.get("Employment") or row.get("Employed") or row.get("Jobs") or row.get("Employed persons")
                if val:
                    try:
                        r["num_jobs_2024"] = str(int(float(str(val).replace(",", ""))))
                        merged += 1
                    except (TypeError, ValueError):
                        pass
        return merged
    try:
        import openpyxl
    except ImportError:
        print("Install openpyxl to merge ABS xlsx: pip install openpyxl")
        return 0
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sheet in wb.worksheets:
        rows_iter = sheet.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            continue
        code_col = find_column(header, ["anzsco", "occupation code", "code", "unit group"])
        emp_col = find_column(header, ["employment", "employed", "jobs", "employed persons"])
        if code_col < 0 or emp_col < 0:
            continue
        for row in rows_iter:
            code = norm_code(row[code_col] if code_col < len(row) else None)
            if not code or code not in rows_by_code:
                continue
            r = rows_by_code[code]
            if row[emp_col] is not None:
                try:
                    r["num_jobs_2024"] = str(int(float(row[emp_col])))
                    merged += 1
                except (TypeError, ValueError):
                    pass
    wb.close()
    return merged


def merge_abs_earnings(rows_by_code):
    """Merge ABS earnings file (xlsx or csv) if present."""
    path = ABS_EARNINGS_CSV if os.path.exists(ABS_EARNINGS_CSV) else (ABS_EARNINGS if os.path.exists(ABS_EARNINGS) else None)
    if not path:
        return 0
    merged = 0
    if path.endswith(".csv"):
        with open(path, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                code = norm_code(row.get("ANZSCO") or row.get("Occupation code") or row.get("Code") or "")
                if not code or code not in rows_by_code:
                    continue
                r = rows_by_code[code]
                weekly = row.get("Median_weekly") or row.get("Weekly earnings") or row.get("Median weekly")
                annual = row.get("Median_annual") or row.get("Annual earnings")
                if annual:
                    try:
                        r["median_pay_annual"] = str(int(float(str(annual).replace(",", "").replace("$", ""))))
                        merged += 1
                    except (TypeError, ValueError):
                        pass
                elif weekly:
                    try:
                        r["median_pay_annual"] = str(int(float(str(weekly).replace(",", "").replace("$", "")) * 52))
                        merged += 1
                    except (TypeError, ValueError):
                        pass
        return merged
    try:
        import openpyxl
    except ImportError:
        print("Install openpyxl to merge ABS xlsx: pip install openpyxl")
        return 0
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sheet in wb.worksheets:
        rows_iter = sheet.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            continue
        code_col = find_column(header, ["anzsco", "occupation code", "code"])
        weekly_col = find_column(header, ["median_weekly", "weekly", "earnings weekly"])
        annual_col = find_column(header, ["median_annual", "annual", "earnings annual"])
        if code_col < 0:
            continue
        for row in rows_iter:
            code = norm_code(row[code_col] if code_col < len(row) else None)
            if not code or code not in rows_by_code:
                continue
            r = rows_by_code[code]
            if annual_col >= 0 and annual_col < len(row) and row[annual_col] is not None:
                try:
                    r["median_pay_annual"] = str(int(float(row[annual_col])))
                    merged += 1
                except (TypeError, ValueError):
                    pass
            elif weekly_col >= 0 and weekly_col < len(row) and row[weekly_col] is not None:
                try:
                    r["median_pay_annual"] = str(int(float(row[weekly_col]) * 52))
                    merged += 1
                except (TypeError, ValueError):
                    pass
    wb.close()
    return merged


def merge_osca_education(rows_by_code):
    """Fill entry_education from OSCA Category Descriptions (Skill Level by unit group). Only fills if currently empty."""
    if not os.path.exists(OSCA_DESCRIPTIONS):
        return 0
    try:
        import openpyxl
    except ImportError:
        return 0
    # Table 1: row 5 = header (Identifier, ..., Skill Level = col 8, 1-based -> index 7)
    wb = openpyxl.load_workbook(OSCA_DESCRIPTIONS, read_only=True, data_only=True)
    if "Table 1" not in wb.sheetnames:
        wb.close()
        return 0
    sh = wb["Table 1"]
    # Build unit_group -> minimum skill level (1 = highest education)
    ug_to_level = {}
    for row in sh.iter_rows(min_row=6, values_only=True):
        if not row or len(row) < 8:
            continue
        ident = str(row[0] or "").strip()
        level_val = row[7]
        if not ident or level_val is None:
            continue
        digits = re.sub(r"\D", "", ident)
        ug = digits[:4] if len(digits) >= 4 else ""
        if not ug:
            continue
        try:
            level = int(float(level_val))
        except (TypeError, ValueError):
            continue
        if 1 <= level <= 5:
            if ug not in ug_to_level or level < ug_to_level[ug]:
                ug_to_level[ug] = level
    wb.close()
    merged = 0
    for code, r in rows_by_code.items():
        if r.get("entry_education"):
            continue
        level = ug_to_level.get(code)
        if level is not None:
            r["entry_education"] = SKILL_LEVEL_TO_AQF.get(level, "")
            if r["entry_education"]:
                merged += 1
    return merged


def merge_jds_profiles(rows_by_code):
    """Merge JSA Occupation profiles from australia/jds (Table_1: Employed, Median weekly earnings, Annual employment growth)."""
    if not os.path.isdir(JDS_DIR):
        return 0
    import glob
    pattern = os.path.join(JDS_DIR, "Occupation profiles data*.xlsx")
    files = glob.glob(pattern)
    if not files:
        return 0
    path = sorted(files)[-1]
    try:
        import openpyxl
    except ImportError:
        print("Install openpyxl to merge JDS xlsx: pip install openpyxl")
        return 0
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Table_1" not in wb.sheetnames:
        wb.close()
        return 0
    sh = wb["Table_1"]
    # Table_1: row 7 = header (1-based), row 8+ = data. Col A=ANZSCO, C=Employed, F=Median weekly $, H=Annual employment growth
    merged = 0
    for row in sh.iter_rows(min_row=8, values_only=True):
        if not row or len(row) < 8:
            continue
        raw = row[0]
        code = norm_code(raw)
        if len(code) != 4 or code not in rows_by_code:
            continue
        # Only use rows that are unit-group level (exactly 4-digit ANZSCO), not 6-digit occupation
        raw_str = re.sub(r"\D", "", str(raw or ""))
        if len(raw_str) != 4:
            continue
        r = rows_by_code[code]
        employed = row[2]
        weekly = row[5]
        growth = row[7]
        if employed is not None:
            try:
                r["num_jobs_2024"] = str(int(float(employed)))
                merged += 1
            except (TypeError, ValueError):
                pass
        if weekly is not None and str(weekly).strip().upper() != "N/A":
            try:
                r["median_pay_annual"] = str(int(round(float(weekly) * 52)))
                merged += 1
            except (TypeError, ValueError):
                pass
        if growth is not None and employed is not None:
            try:
                emp_val = float(employed)
                if emp_val != 0:
                    val_2025 = str(int(round(float(growth) / emp_val * 100)))
                    r["outlook_pct"] = val_2025
                    r["outlook_pct_2025_actual"] = val_2025  # keep 2025 actual before Table_6 overwrites outlook_pct
                    merged += 1
            except (TypeError, ValueError):
                pass
    wb.close()
    return merged


def merge_jds_projections_table6(rows_by_code):
    """Overwrite outlook_pct with 10-year (May 2025 → May 2035) % from JSA Employment Projections Table_6."""
    if not os.path.isdir(JDS_PROJECTIONS_DIR):
        return 0
    import glob
    pattern = os.path.join(JDS_PROJECTIONS_DIR, "employment_projections*.xlsx")
    files = glob.glob(pattern)
    if not files:
        return 0
    path = sorted(files)[-1]
    try:
        import openpyxl
    except ImportError:
        return 0
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet_name = "Table_6 Occupation Unit Group"
    if sheet_name not in wb.sheetnames:
        wb.close()
        return 0
    sh = wb[sheet_name]
    # Row 10+: col B=NFD, C=ANZSCO, L=10-year % (decimal, e.g. 0.14 = 14%)
    merged = 0
    for row in sh.iter_rows(min_row=10, values_only=True):
        if not row or len(row) < 12:
            continue
        nfd = row[1]
        raw = row[2]
        code = norm_code(raw)
        if code not in rows_by_code or (nfd and str(nfd).strip().upper() != "N"):
            continue
        raw_str = re.sub(r"\D", "", str(raw or ""))
        if len(raw_str) != 4:
            continue
        ten_yr_pct = row[11]
        if ten_yr_pct is None:
            continue
        try:
            pct_val = float(ten_yr_pct)
            r = rows_by_code[code]
            r["outlook_pct"] = str(int(round(pct_val * 100)))
            merged += 1
        except (TypeError, ValueError):
            pass
    wb.close()
    return merged


def apply_imputed(rows_by_code):
    """Fill empty outlook/education/pay from data_abs/imputed.json (LLM-imputed). See scripts/impute_missing_with_llm.py."""
    path = os.path.join(DATA_ABS, "imputed.json")
    if not os.path.exists(path):
        return 0
    with open(path, encoding="utf-8") as f:
        imputed = json.load(f)
    merged = 0
    for r in rows_by_code.values():
        slug = slugify(r.get("title", ""))
        if not slug:
            continue
        if not r.get("outlook_pct") and slug in imputed.get("outlook", {}):
            r["outlook_pct"] = str(imputed["outlook"][slug]["outlook_pct"])
            merged += 1
        if not r.get("entry_education") and slug in imputed.get("education", {}):
            r["entry_education"] = imputed["education"][slug]["entry_education"]
            merged += 1
        if not r.get("median_pay_annual") and slug in imputed.get("pay", {}):
            r["median_pay_annual"] = str(imputed["pay"][slug]["median_pay_annual"])
            merged += 1
    return merged


def apply_ai_forecast(rows_by_code):
    """Fill outlook_pct_ai from data_abs/ai_forecast_outlook.json (from scripts/ai_forecast_au.py)."""
    path = os.path.join(DATA_ABS, "ai_forecast_outlook.json")
    if not os.path.exists(path):
        return 0
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    merged = 0
    # Expect list of { "anzsco_code": "1111", "adjusted_growth_pct_2035": 5 } or dict by code
    if isinstance(data, dict):
        by_code = data
    else:
        by_code = {}
        for item in data:
            code = (item.get("anzsco_code") or "").strip()
            if code:
                by_code[code] = item
    for code, r in rows_by_code.items():
        item = by_code.get(code)
        if item is None:
            continue
        pct = item.get("adjusted_growth_pct_2035")
        if pct is not None:
            try:
                r["outlook_pct_ai"] = str(int(round(float(pct))))
                merged += 1
            except (TypeError, ValueError):
                pass
    return merged


def main():
    if not os.path.exists(COMBINED_CSV):
        print("Run build_occupations_from_anzsco.py first to create occupations_combined.csv")
        return

    with open(COMBINED_CSV, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fieldnames = list(reader.fieldnames)
        if "outlook_pct_2025_actual" not in fieldnames:
            fieldnames.append("outlook_pct_2025_actual")
        if "outlook_pct_ai" not in fieldnames:
            fieldnames.append("outlook_pct_ai")
        rows = list(reader)

    rows_by_code = {}
    for r in rows:
        code = norm_code(r.get("anzsco_code") or "")
        if code:
            rows_by_code[code] = r

    # JDS profiles first (employment, pay, past-year outlook); then Projections Table_6 overwrites outlook with 2035 10-yr %
    n_jds = merge_jds_profiles(rows_by_code)
    n_jsa = merge_jsa(rows_by_code)
    n_proj = merge_jds_projections_table6(rows_by_code)
    n_abs_emp = merge_abs_employment(rows_by_code)
    n_abs_earn = merge_abs_earnings(rows_by_code)
    n_osca = merge_osca_education(rows_by_code)
    n_imputed = apply_imputed(rows_by_code)
    n_ai = apply_ai_forecast(rows_by_code)

    with open(COMBINED_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    print(f"Merged real data: JDS profiles {n_jds}, JSA {n_jsa}, Projections 2035 (Table_6) {n_proj}, ABS emp {n_abs_emp}, ABS earn {n_abs_earn}, OSCA {n_osca}, imputed {n_imputed}, AI forecast {n_ai}.")
    print("Next: run make_csv_au.py then build_site_data_au.py.")


if __name__ == "__main__":
    main()
