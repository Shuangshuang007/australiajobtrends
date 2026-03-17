"""
Step 1 of AI Forecast: build a rough AU occupation → US BLS occupation mapping using GPT-4.1.

Reads:
  - data_abs/occupations_combined.csv (AU: title, anzsco_code)
  - ../australia/reference/USA_occupation.xlsx Table 1.2 (US: SOC, title, employment change % 2024-34)

Calls OpenAI (or OpenRouter) with model gpt-4.1 to match each AU occupation to the best US occupation.
Writes: ../australia/reference/au_us_occupation_mapping.csv

Usage (from Job Outlook Australia folder):
  python scripts/map_au_us_occupations.py
  python scripts/map_au_us_occupations.py --start 0 --end 50   # first 50 AU occupations
  python scripts/map_au_us_occupations.py --model gpt-4o      # fallback if gpt-4.1 not available
"""

import csv
import json
import os
import re
import time
import argparse

try:
    import httpx
except ImportError:
    httpx = None

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.dirname(SCRIPT_DIR)
DATA_ABS = os.path.join(BASE_DIR, "data_abs")
AUSTRALIA_DIR = os.path.join(os.path.dirname(BASE_DIR), "australia")
REFERENCE_DIR = os.path.join(AUSTRALIA_DIR, "reference")
USA_XLSX = os.path.join(REFERENCE_DIR, "USA_occupation.xlsx")
COMBINED_CSV = os.path.join(DATA_ABS, "occupations_combined.csv")
OUTPUT_CSV = os.path.join(REFERENCE_DIR, "au_us_occupation_mapping.csv")

# AI_FORECAST_DESIGN.md: mapping uses gpt-4.1
DEFAULT_MODEL = "gpt-4.1"
OPENAI_API_URL = "https://api.openai.com/v1/chat/completions"
BATCH_SIZE = 25


def load_au_occupations():
    """Load AU occupations: anzsco_code, title."""
    if not os.path.exists(COMBINED_CSV):
        return []
    rows = []
    with open(COMBINED_CSV, newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            code = (r.get("anzsco_code") or "").strip()
            title = (r.get("title") or "").strip()
            if code and title:
                rows.append({"anzsco_code": code, "au_title": title})
    return rows


def load_us_occupations():
    """Load US from USA_occupation.xlsx Table 1.2, Line item only. Returns list of (soc, title, pct)."""
    if not os.path.exists(USA_XLSX):
        return []
    try:
        import openpyxl
    except ImportError:
        return []
    wb = openpyxl.load_workbook(USA_XLSX, read_only=True, data_only=True)
    sheet_name = "Table 1.2"
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    sh = wb[sheet_name]
    # Row 3+: A=title, B=code, C=type, I=employment change percent (index 8)
    out = []
    for row in sh.iter_rows(min_row=3, values_only=True):
        if not row or len(row) < 9:
            continue
        otype = row[2]
        if otype != "Line item":
            continue
        code = (row[1] or "").strip()
        title = (row[0] or "").strip()
        if not code or code == "00-0000":
            continue
        pct = row[8]
        try:
            pct_val = float(pct) if pct is not None else None
        except (TypeError, ValueError):
            pct_val = None
        out.append({"soc": code, "title": title, "projection_pct": pct_val})
    wb.close()
    return out


def build_prompt(au_batch, us_list_text):
    return f"""You are matching Australian occupations (ANZSCO) to the closest US BLS occupation (SOC) for labour market comparison. Australia often lags the US by 2–3 years; US projections are a leading indicator.

US occupations (SOC, title, employment change % 2024–34):
{us_list_text}

Australian occupations to match (anzsco_code, title):
{json.dumps(au_batch, ensure_ascii=False, indent=2)}

For each Australian occupation, pick the single best US match (or "no_match" if none is close). Consider job content and title similarity, not just wording.

Respond with ONLY a JSON array, one object per AU occupation, in order. No other text.
[
  {{ "anzsco_code": "...", "au_title": "...", "us_soc": "XX-XXXX" or null, "us_title": "..." or null, "us_projection_pct": number or null, "match_confidence": "high" | "medium" | "low" | "none" }},
  ...
]"""


def call_llm(client, api_key, model, au_batch, us_list_text):
    """Call OpenAI-compatible API; returns list of mapping dicts."""
    prompt = build_prompt(au_batch, us_list_text)
    url = os.environ.get("OPENROUTER_API_URL", OPENAI_API_URL)
    if "openrouter" in url.lower():
        headers = {"Authorization": f"Bearer {api_key}", "HTTP-Referer": "https://github.com"}
    else:
        headers = {"Authorization": f"Bearer {api_key}"}
    resp = client.post(
        url,
        headers=headers,
        json={
            "model": model,
            "messages": [
                {"role": "user", "content": prompt},
            ],
            "temperature": 0.1,
        },
        timeout=120,
    )
    resp.raise_for_status()
    content = resp.json()["choices"][0]["message"]["content"].strip()
    # Strip markdown code block if present
    if content.startswith("```"):
        content = re.sub(r"^```\w*\n?", "", content)
        content = re.sub(r"\n?```\s*$", "", content)
    return json.loads(content)


def main():
    try:
        from dotenv import load_dotenv
        load_dotenv()
        load_dotenv(os.path.join(BASE_DIR, ".env.local"))
        load_dotenv(os.path.join(os.path.dirname(BASE_DIR), ".env.local"))
    except ImportError:
        pass
    parser = argparse.ArgumentParser(description="Map AU occupations to US BLS (GPT-4.1)")
    parser.add_argument("--model", default=DEFAULT_MODEL, help=f"Model for mapping (default: {DEFAULT_MODEL})")
    parser.add_argument("--start", type=int, default=0)
    parser.add_argument("--end", type=int, default=None)
    parser.add_argument("--batch-size", type=int, default=BATCH_SIZE)
    parser.add_argument("--delay", type=float, default=1.0)
    parser.add_argument("--dry-run", action="store_true", help="Print prompt only, no API call")
    args = parser.parse_args()

    if not httpx:
        print("Install httpx: pip install httpx")
        return

    api_key = os.environ.get("OPENAI_API_KEY") or os.environ.get("OPENROUTER_API_KEY")
    if not api_key and not args.dry_run:
        print("Set OPENAI_API_KEY or OPENROUTER_API_KEY (e.g. in .env.local)")
        return

    au_list = load_au_occupations()
    us_list = load_us_occupations()
    if not au_list:
        print(f"No AU occupations in {COMBINED_CSV}. Run merge_real_data.py first.")
        return
    if not us_list:
        print(f"No US occupations in {USA_XLSX} Table 1.2.")
        return

    us_list_text = "\n".join(
        f"  {u['soc']} | {u['title'][:50]} | {u['projection_pct']}" for u in us_list
    )
    # Truncate if huge (keep first 400 lines for context)
    if us_list_text.count("\n") > 400:
        us_list_text = "\n".join(us_list_text.split("\n")[:400]) + "\n  ... (more)"

    subset = au_list[args.start:args.end]
    existing = {}
    if os.path.exists(OUTPUT_CSV):
        with open(OUTPUT_CSV, newline="", encoding="utf-8") as f:
            for r in csv.DictReader(f):
                # Only skip re-mapping if we already have a US match
                if (r.get("us_soc") or "").strip():
                    existing[r["anzsco_code"]] = r

    os.makedirs(REFERENCE_DIR, exist_ok=True)
    fieldnames = ["anzsco_code", "au_title", "us_soc", "us_title", "us_projection_pct", "match_confidence"]

    start_idx = args.start
    for i in range(0, len(subset), args.batch_size):
        batch = subset[i : i + args.batch_size]
        batch_codes = {b["anzsco_code"] for b in batch}
        if all(c in existing for c in batch_codes):
            start_idx += len(batch)
            continue
        if args.dry_run:
            print("Prompt (first 2000 chars):", build_prompt(batch, us_list_text)[:2000])
            return
        client = httpx.Client()
        try:
            result = call_llm(client, api_key, args.model, batch, us_list_text)
        finally:
            client.close()
        if not isinstance(result, list):
            result = [result]
        for m in result:
            code = m.get("anzsco_code") or ""
            if not code:
                continue
            existing[code] = {
                "anzsco_code": code,
                "au_title": m.get("au_title") or "",
                "us_soc": m.get("us_soc") or "" if m.get("us_soc") is not None else "",
                "us_title": m.get("us_title") or "",
                "us_projection_pct": m.get("us_projection_pct") if m.get("us_projection_pct") is not None else "",
                "match_confidence": m.get("match_confidence") or "",
            }
        start_idx += len(batch)
        print(f"Mapped batch {i // args.batch_size + 1}: {len(batch)} occupations, total mapped {len(existing)}")
        time.sleep(args.delay)

    # Write one row per AU occupation; use existing mapping or no-match row
    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for occ in au_list:
            code = occ["anzsco_code"]
            row = existing.get(code)
            if not row:
                row = {"anzsco_code": code, "au_title": occ["au_title"], "us_soc": "", "us_title": "", "us_projection_pct": "", "match_confidence": "none"}
            w.writerow({k: row.get(k, "") for k in fieldnames})

    print(f"Wrote {OUTPUT_CSV} with {len(au_list)} rows.")


if __name__ == "__main__":
    main()
